# DEBUG SCRIPT: Add this to test your panel data fetching
# Save as: test_panel_data.py

import os
import sys
import json
import logging
from datetime import datetime

# Add the app directory to the path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.services.grafana_api import GrafanaService
from app.services.report_service import ReportService
from app.config import settings

# Setup logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def test_panel_data():
    """Test fetching panel data and generating a report"""
    
    logger.info("=" * 80)
    logger.info("TESTING PANEL DATA FETCHING")
    logger.info("=" * 80)
    
    # Initialize services
    grafana_service = GrafanaService()
    report_service = ReportService(reports_dir=settings.REPORTS_DIR)
    
    # Step 1: Get dashboards
    logger.info("\n[STEP 1] Fetching dashboards...")
    try:
        dashboards = grafana_service.get_dashboards()
        logger.info(f"Found {len(dashboards)} dashboards")
        for dash in dashboards[:5]:  # Show first 5
            logger.info(f"  - {dash['title']} (UID: {dash['uid']})")
    except Exception as e:
        logger.error(f"Failed to get dashboards: {e}")
        return
    
    if not dashboards:
        logger.error("No dashboards found!")
        return
    
    # Step 2: Get panels from first dashboard
    dashboard_uid = dashboards[0]['uid']
    logger.info(f"\n[STEP 2] Fetching panels from dashboard: {dashboards[0]['title']}")
    
    try:
        panels = grafana_service.get_dashboard_panels(dashboard_uid)
        logger.info(f"Found {len(panels)} panels")
        for panel in panels[:5]:  # Show first 5
            logger.info(f"  - Panel {panel['id']}: {panel['title']} (type: {panel['type']})")
    except Exception as e:
        logger.error(f"Failed to get panels: {e}")
        return
    
    if not panels:
        logger.error("No panels found!")
        return
    
    # Step 3: Fetch data from first panel
    panel_id = panels[0]['id']
    logger.info(f"\n[STEP 3] Fetching data from panel {panel_id}: {panels[0]['title']}")
    
    time_range = {
        "from": "now-24h",
        "to": "now"
    }
    
    try:
        panel_data = grafana_service.get_panel_data(dashboard_uid, panel_id, time_range)
        
        # Log the results
        logger.info(f"\n[RESULTS]")
        logger.info(f"  Panel ID: {panel_data['panel']['id']}")
        logger.info(f"  Panel Title: {panel_data['panel']['title']}")
        logger.info(f"  Panel Type: {panel_data['panel']['type']}")
        logger.info(f"  Fields: {panel_data['fields']}")
        logger.info(f"  Number of rows: {len(panel_data['rows'])}")
        
        if panel_data['rows']:
            logger.info(f"  First row: {panel_data['rows'][0]}")
        else:
            logger.warning("  ⚠️  NO ROWS RETURNED!")
            
        # Check for errors
        if 'error' in panel_data['panel']:
            logger.error(f"  ❌ Panel Error: {panel_data['panel']['error']}")
            
    except Exception as e:
        logger.error(f"Failed to get panel data: {e}", exc_info=True)
        return
    
    # Step 4: Try to generate a test report
    logger.info(f"\n[STEP 4] Generating test report...")
    
    try:
        panels_data = [panel_data]
        
        filename = f"test_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        report_title = "Test Security Report"
        
        file_path = report_service.generate_xlsx_from_panels(
            panels_data,
            filename,
            report_title,
            time_range,
            logo_path=None,
            company_name="Test Company"
        )
        
        logger.info(f"✅ Report generated successfully!")
        logger.info(f"  File: {file_path}")
        logger.info(f"  File size: {os.path.getsize(file_path)} bytes")
        
        # Check if file has actual data
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        logger.info(f"  Sheets: {wb.sheetnames}")
        
        if len(wb.sheetnames) > 1:
            ws = wb[wb.sheetnames[1]]  # First data sheet
            logger.info(f"  Data sheet rows: {ws.max_row}")
            logger.info(f"  Data sheet cols: {ws.max_column}")
            
            if ws.max_row < 5:
                logger.warning("  ⚠️  Very few rows in the data sheet!")
        
    except Exception as e:
        logger.error(f"Failed to generate report: {e}", exc_info=True)
        return
    
    logger.info("\n" + "=" * 80)
    logger.info("TEST COMPLETED")
    logger.info("=" * 80)

if __name__ == "__main__":
    test_panel_data()