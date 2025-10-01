# report_service.py
import os
import json
import logging
from typing import Any, Dict, List, Optional
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

# Default date format for timestamps
DEFAULT_DATE_FORMAT = "%Y-%m-%d %H:%M:%S"


class ReportService:
    def __init__(self, reports_dir: str, logger: Optional[logging.Logger] = None) -> None:
        self.reports_dir = reports_dir
        os.makedirs(self.reports_dir, exist_ok=True)
        self.logger = logger or logging.getLogger(__name__)

    # ---------------------------
    # Public: single-sheet export
    # ---------------------------
    def generate_xlsx(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        fields: List[str],
        report_title: str,
        logo_path: Optional[str] = None,
    ) -> str:
        """Generate an XLSX report with the given data, fields, and formatting."""
        if not filename.endswith(".xlsx"):
            filename = f"{filename}.xlsx"
        file_path = os.path.join(self.reports_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Security Report"

        # Styles
        title_font = Font(size=16, bold=True)
        header_font = Font(size=12, bold=True)
        small_italic = Font(size=10, italic=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        # Header area
        current_row = 1

        # Logo
        if logo_path and os.path.exists(logo_path.lstrip("/")):
            try:
                img = Image(logo_path.lstrip("/"))
                img = self._resize_logo(img, max_width=200, max_height=80)
                ws.add_image(img, "A1")
                current_row += 5
            except Exception:
                pass

        # Title
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max(1, len(fields)))
        cell = ws.cell(row=current_row, column=1, value=report_title or "Security Report")
        cell.font = title_font
        current_row += 1

        # Timestamp
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max(1, len(fields)))
        ts = datetime.now().strftime(DEFAULT_DATE_FORMAT)
        tcell = ws.cell(row=current_row, column=1, value=f"Generated: {ts}")
        tcell.font = small_italic
        current_row += 2  # blank line before table

        # Table header
        header_row = current_row
        safe_headers = [self._safe_header(h) for h in (fields or [])]
        if not safe_headers:
            if data:
                safe_headers = [self._safe_header(k) for k in data[0].keys()]
            else:
                safe_headers = ["Column 1"]

        for idx, title in enumerate(safe_headers, start=1):
            c = ws.cell(row=header_row, column=idx, value=title)
            c.font = header_font
            c.alignment = center_align
            c.fill = header_fill

        # Data rows
        data_start = header_row + 1
        r = data_start
        for row in data or []:
            for idx, key in enumerate(safe_headers, start=1):
                value = row.get(key, row.get(key.lower(), ""))
                ws.cell(row=r, column=idx, value=self._display_value(value, guess_from_header=key))
            r += 1

        # Freeze panes
        ws.freeze_panes = f"A{data_start}"

        # Excel table
        if r - 1 >= header_row:
            try:
                ref = f"A{header_row}:{get_column_letter(len(safe_headers))}{max(header_row, r-1)}"
                tbl = Table(displayName="SecurityReportTable", ref=ref)
                tbl.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                ws.add_table(tbl)
            except Exception:
                pass

        self._autosize_worksheet(ws, header_row=header_row, last_row=r - 1, max_width=60)
        wb.save(file_path)
        return file_path

    # ---------------------------------
    # Public: multi-sheet export
    # ---------------------------------
    def generate_xlsx_from_panels(
        self,
        panels_data: List[Dict[str, Any]],
        filename: str,
        report_title: str,
        time_range: Dict[str, str],
        logo_path: Optional[str] = None,
        company_name: Optional[str] = None,
    ) -> str:
        """Generate an XLSX report with multiple sheets, one per panel."""
        if not filename.endswith(".xlsx"):
            filename = f"{filename}.xlsx"
        file_path = os.path.join(self.reports_dir, filename)

        wb = Workbook()
        cover = wb.active
        cover.title = "Summary"

        # Styles
        title_font = Font(size=16, bold=True)
        subtitle_font = Font(size=14, bold=True)
        header_font = Font(size=12, bold=True)
        normal_font = Font(size=11)
        small_font = Font(size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        # --- Cover sheet ---
        row = 1
        if logo_path and os.path.exists(logo_path.lstrip("/")):
            try:
                img = Image(logo_path.lstrip("/"))
                img = self._resize_logo(img, max_width=200, max_height=80)
                cover.add_image(img, "A1")
                row += 5
            except Exception:
                pass

        cover.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        tcell = cover.cell(row=row, column=1, value=report_title or "Security Report")
        tcell.font = title_font
        row += 2

        if company_name:
            cover.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ccell = cover.cell(row=row, column=1, value=f"Company: {company_name}")
            ccell.font = header_font
            row += 1

        if time_range.get("from") and time_range.get("to"):
            cover.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            trcell = cover.cell(
                row=row,
                column=1,
                value=f"Time Range: {time_range['from']} to {time_range['to']}"
            )
            trcell.font = header_font
            row += 1

        ts = datetime.now().strftime(DEFAULT_DATE_FORMAT)
        cover.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        gcell = cover.cell(row=row, column=1, value=f"Generated: {ts}")
        gcell.font = small_font
        gcell.alignment = left_align
        row += 2

        # TOC
        cover.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        toch = cover.cell(row=row, column=1, value="Report Contents")
        toch.font = subtitle_font
        row += 1

        for col_idx, hdr in enumerate(["Sheet", "Title", "Type", "Description"], start=1):
            c = cover.cell(row=row, column=col_idx, value=hdr)
            c.font = header_font
            c.fill = subheader_fill
            c.alignment = center_align
            c.border = thin_border
        row += 1

        # --- Per-panel sheets ---
        for i, pdata in enumerate(panels_data or [], start=1):
            panel_info = pdata.get("panel", {})
            p_title = panel_info.get("title") or f"Panel {i}"
            p_type = panel_info.get("type") or "unknown"
            p_desc = panel_info.get("description") or ""
            sheet_name = f"Sheet{i}"

            for value, col in zip([sheet_name, p_title, p_type, p_desc], range(1, 5)):
                c = cover.cell(row=row, column=col, value=value)
                c.font = normal_font
                c.alignment = left_align
                c.border = thin_border
            row += 1

            ws = wb.create_sheet(title=sheet_name)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
            ws.cell(row=1, column=1, value=p_title).font = title_font

            if p_desc:
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
                ws.cell(row=2, column=1, value=p_desc).font = normal_font

            fields = [self._safe_header(h) for h in pdata.get("fields", [])]
            rows = pdata.get("rows", [])

            if fields and rows:
                header_row = 4
                for col_idx, hdr in enumerate(fields, start=1):
                    c = ws.cell(row=header_row, column=col_idx, value=hdr)
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = center_align

                data_start = header_row + 1
                r = data_start
                for row_vals in rows:
                    for col_idx, value in enumerate(row_vals[: len(fields)], start=1):
                        ws.cell(row=r, column=col_idx,
                                value=self._display_value(value, guess_from_header=fields[col_idx-1]))
                    r += 1

                ws.freeze_panes = f"A{data_start}"
                self._autosize_worksheet(ws, header_row=header_row, last_row=r - 1, max_width=60)

        wb.save(file_path)
        return file_path

    # ---------------------------
    # Helpers
    # ---------------------------
    def _resize_logo(self, img: Image, max_width: int, max_height: int) -> Image:
        try:
            h, w = img.height, img.width
            if h > max_height:
                scale = max_height / h
                img.height = max_height
                img.width = int(w * scale)
            if img.width > max_width:
                scale = max_width / img.width
                img.width = max_width
                img.height = int(h * scale)
        except Exception:
            pass
        return img

    def _safe_header(self, field: Any) -> str:
        return "" if field is None else str(field)

    def _display_value(self, value: Any, guess_from_header: str = "") -> Any:
        if isinstance(value, (list, tuple)):
            return ", ".join(str(v) for v in value)
        if isinstance(value, dict):
            return json.dumps(value, ensure_ascii=False)
        return value

    def _autosize_worksheet(self, ws, header_row: int, last_row: int,
                            max_width: int = 60, min_width: int = 10) -> None:
        try:
            if last_row < header_row:
                return
            for col_idx in range(1, ws.max_column + 1):
                letter = get_column_letter(col_idx)
                max_len = 0
                hval = ws.cell(row=header_row, column=col_idx).value
                if hval:
                    max_len = max(max_len, len(str(hval)))
                for r in range(header_row + 1, last_row + 1):
                    val = ws.cell(row=r, column=col_idx).value
                    if val:
                        max_len = max(max_len, min(len(str(val)), max_width))
                ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))
        except Exception:
            pass
